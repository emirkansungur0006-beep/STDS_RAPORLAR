import openpyxl
import os

path = r'C:\Users\EMİRKAN SUNGUR\Desktop\STDS_RAPORLAR\RAPORLAR\Gelişim_Planı_İller\ADANA.xlsx'

try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n--- SHEET: {sn} ---")
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20)):
            row_vals = [str(c).strip() for c in row if c is not None]
            if any(v for v in row_vals if len(v) > 2):
                print(f"R{i+1}: {row_vals}")
    wb.close()
except Exception as e:
    print(f"Error: {e}")
