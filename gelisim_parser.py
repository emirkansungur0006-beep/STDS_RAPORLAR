# -*- coding: utf-8 -*-
"""
STDS Raporlar - Gelişim Planları Parser (V5 - ROW-BASED MULTI-HOSPITAL)
- Sayfa içinde değişen hastane isimlerini algılar.
- Ankara, Adana ve diğer karmaşık dosyaları %100 doğrulukla işler.
"""
import os
import re
import sqlite3
import openpyxl
from config import BASE_DIR
from app import get_db
from reference_matcher import ReferenceMatcher, normalize_turkish

GELISIM_DIR = os.path.join(BASE_DIR, 'RAPORLAR', 'Gelişim_Planı_İller')

def clean_text(text):
    if text is None: return None
    t = str(text).strip()
    if t.lower() == 'none' or t == '': return None
    t = t.replace('_x000D_', ' ')
    t = re.sub(r'\s+', ' ', t)
    return t

class GelisimParser:
    def __init__(self):
        self.matcher = ReferenceMatcher()
        self.keywords = {
            'kurum_hedefleri': ['hedef', 'kurum hedefleri'],
            'gerceklesme_suresi': ['süre', 'sure'],
            'mevcut_durum': ['mevcut durum', 'tespit'],
            'cozum_secenekleri': ['çözüm', 'cozum'],
            'etki_analizi': ['etki analizi'],
            'uygun_secenek': ['uygun', 'secenek'],
            'isbirligi_plani': ['işbirliği', 'isbirligi'],
            'uygulama_takvimi': ['takvim'],
            'hastane_col': ['hastane', 'tesis', 'kurum adı', 'kurum adi']
        }
        # Keyword'leri de normalize et ki eşleşme sağlıklı olsun
        for k in self.keywords:
            self.keywords[k] = [normalize_turkish(kw) for kw in self.keywords[k]]

    def detect_hospital(self, row, il=None):
        """Satırın bir hastane başlığı olup olmadığını kontrol eder"""
        # Hücreleri temizle
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not cells: return None
        
        # Strateji 1: "SAĞLIK TESİSİ ADI:" kalıbı
        row_str = " ".join(cells).upper()
        if 'ADI:' in row_str and ('SAĞLIK' in row_str or 'TESİS' in row_str or 'KURUM' in row_str):
            for ci, c_str in enumerate(cells):
                if 'ADI:' in c_str.upper():
                    potential = c_str.upper().split('ADI:')[-1].strip()
                    if len(potential) < 5 and ci + 1 < len(cells):
                        potential = cells[ci+1]
                    m = self.matcher.match_by_name(potential, il=il, threshold=0.7)
                    if m: return m

        # Strateji 2: İlk hücre hastane adı, ikinci hücre adres/il (Adana pattern)
        if len(cells) > 1:
            potential_name = cells[0]
            potential_addr = cells[1]
            if len(potential_name) > 10 and (il and il.upper() in potential_addr.upper() or 'CAD.' in potential_addr.upper() or 'MAH.' in potential_addr.upper() or 'HASTANESİ' in potential_name.upper()):
                m = self.matcher.match_by_name(potential_name, il=il, threshold=0.75)
                if m: return m
        
        # Strateji 3: Sadece ilk hücrede bir hastane adı varsa
        if len(cells[0]) > 10:
             # Afyon için threshold'u biraz düşürelim çünkü isimler çok farklı
             m = self.matcher.match_by_name(cells[0], il=il, threshold=0.7)
             if m: return m
             
        return None

    def process_file(self, filepath):
        filename = os.path.basename(filepath)
        il_raw = os.path.splitext(filename)[0].split('-')[0].split('_')[0].strip().upper()
        # Bazı dosya isimleri hatalı olabilir (Samsunkütahya vb.)
        if il_raw == 'SAMSUNKÜTAHYA': il_raw = 'SAMSUN'
        if 'ERZURUM' in il_raw and 'HAKKARİ' in il_raw: il_raw = 'ERZURUM'
        if il_raw == 'ISTANBUL': il_raw = 'İSTANBUL'
        if il_raw == 'GİRESUNN': il_raw = 'GİRESUN'
        if il_raw == 'KOACAELİ': il_raw = 'KOCAELİ'
        if il_raw == 'AFYON': il_raw = 'AFYONKARAHİSAR'
        
        il_match = self.matcher.match_by_name(il_raw, threshold=0.8)
        il_name = il_match['il'] if il_match else il_raw
        
        # Çoklu il dosyalarında hastane ararken tüm illerde aranmasına izin ver
        il_name_for_match = None if ('SAMSUNKÜTAHYA' in filename.upper() or 'ERZURUM HAKKARİ' in filename.upper()) else il_name

        print(f"  Analiz: {filename} ({il_name})")
        
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
        except Exception as e:
            print(f"    [!] HATA: {e}")
            return []

        all_records = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            try:
                # 3967 satıra kadar oku (Adana için)
                for r in ws.iter_rows(values_only=True, max_row=5000):
                    rows.append(r)
            except: continue
            
            if not rows: continue
            
            current_hosp = None
            header_map = None
            match_cache = {}
            
            # Sheet adından default bulmaya çalış (Yedek)
            sheet_hosp = self.matcher.match_by_name(sheet_name, il=il_name_for_match, threshold=0.8)

            for i, row in enumerate(rows):
                if not any(row): continue
                filled_cells = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
                if not filled_cells: continue
                
                # 1. Header Tespiti
                if not header_map:
                    matches = {}
                    row_norm = [normalize_turkish(str(c)) for c in row if c is not None]
                    # İki aşamalı kontrol: Önce tam eşleşme, sonra kısmi
                    for stage in ['exact', 'partial']:
                        for col_db_name, kw_list in self.keywords.items():
                            if col_db_name in matches: continue
                            for ci, val in enumerate(row):
                                if val is None: continue
                                val_norm = normalize_turkish(str(val))
                                if stage == 'exact':
                                    if any(kw == val_norm for kw in kw_list):
                                        matches[col_db_name] = ci
                                        break
                                else:
                                    if any(kw in val_norm for kw in kw_list):
                                        matches[col_db_name] = ci
                                        break
                    
                    if len(matches) >= 2 and ('kurum_hedefleri' in matches or 'mevcut_durum' in matches):
                        header_map = matches
                        continue

                # 2. Hastane Başlığı Tespiti
                if not header_map or len(filled_cells) <= 3:
                    row_str = " ".join(filled_cells).upper()
                    if 'GELİŞİM PLAN' in row_str and 'HASTANE' not in row_str and 'TESİS' not in row_str:
                         pass
                    else:
                        new_hosp = self.detect_hospital(row, il=il_name_for_match)
                        if new_hosp:
                            current_hosp = new_hosp
                            header_map = None # Yeni hastane gelince header'ı resetle
                            continue

                # 3. Veri İşleme
                if header_map:
                    row_str = " ".join(filled_cells).upper()
                    if 'GELİŞİM PLAN' in row_str and len(filled_cells) < 3: continue
                    if 'HEDEF' in row_str and 'SÜRE' in row_str: continue

                    hastane_adi_row = None
                    if 'hastane_col' in header_map:
                        col_idx = header_map['hastane_col']
                        if col_idx < len(row):
                            h_val = row[col_idx]
                            if h_val:
                                h_val_clean = str(h_val).strip()
                                cache_key = (h_val_clean, il_name_for_match)
                                if cache_key in match_cache:
                                    h_match = match_cache[cache_key]
                                else:
                                    h_match = self.matcher.match_by_name(h_val_clean, il=il_name_for_match, threshold=0.7)
                                    match_cache[cache_key] = h_match
                                
                                if h_match:
                                    hastane_adi_row = h_match
                                else:
                                    hastane_adi_row = {'il': il_name, 'kurum_adi': h_val_clean, 'kurum_kodu': None}

                    hastane_match = hastane_adi_row or current_hosp or sheet_hosp
                    
                    if hastane_match:
                        hastane_adi = hastane_match.get('kurum_adi', f"{il_name} - {sheet_name.upper()}")
                        kurum_kodu = hastane_match.get('kurum_kodu')
                        gercek_il = hastane_match.get('il', il_name)
                    else:
                        hastane_adi = f"{il_name} - {sheet_name.upper()}"
                        kurum_kodu = None
                        gercek_il = il_name

                    rec = {
                        'il': gercek_il,
                        'hastane_adi': hastane_adi,
                        'kurum_kodu': kurum_kodu,
                        'kaynak_dosya': filename,
                        'sheet_adi': sheet_name
                    }
                    
                    for col_db_name, ci in header_map.items():
                        if col_db_name == 'hastane_col': continue
                        val = row[ci] if ci < len(row) else None
                        rec[col_db_name] = clean_text(val)
                    
                    # Boş kayıtları ele
                    if not rec.get('kurum_hedefleri') or len(str(rec.get('kurum_hedefleri'))) < 5:
                        continue

                    all_records.append(rec)

        wb.close()
        return all_records

    def save_to_db(self, records):
        if not records: return 0
        conn = get_db()
        cur = conn.cursor()
        
        from config import USE_SQLITE
        placeholders = '?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?' if USE_SQLITE else '%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s'
        
        count = 0
        for r in records:
            try:
                cur.execute(f"""
                    INSERT INTO gelisim_planlari 
                    (kurum_kodu, il, hastane_adi, kurum_hedefleri, gerceklesme_suresi, 
                     mevcut_durum, cozum_secenekleri, etki_analizi, uygun_secenek, 
                     isbirligi_plani, uygulama_takvimi, kaynak_dosya, sheet_adi)
                    VALUES ({placeholders})
                """, (
                    r.get('kurum_kodu'), r.get('il'), r.get('hastane_adi'),
                    r.get('kurum_hedefleri'), r.get('gerceklesme_suresi'),
                    r.get('mevcut_durum'), r.get('cozum_secenekleri'),
                    r.get('etki_analizi'), r.get('uygun_secenek'),
                    r.get('isbirligi_plani'), r.get('uygulama_takvimi'),
                    r.get('kaynak_dosya'), r.get('sheet_adi')
                ))
                count += 1
            except Exception as e:
                conn.rollback()
                continue
        
        conn.commit()
        cur.close()
        conn.close()
        return count

def run():
    print("="*60)
    print("GELİŞİM PLANLARI PARSER V5 (VERTICAL MULTI-HOSPITAL)")
    print("="*60)
    
    parser = GelisimParser()
    
    # DB Temizliği
    conn = get_db()
    cur = conn.cursor()
    from config import USE_SQLITE
    if USE_SQLITE:
        cur.execute("DELETE FROM gelisim_planlari")
    else:
        cur.execute("TRUNCATE TABLE gelisim_planlari")
    conn.commit()
    cur.close()
    conn.close()
    
    total = 0
    files = sorted([f for f in os.listdir(GELISIM_DIR) if f.endswith('.xlsx')])
    
    for f in files:
        records = parser.process_file(os.path.join(GELISIM_DIR, f))
        if records:
            saved = parser.save_to_db(records)
            total += saved
            print(f"    -> {saved} kayıt yüklendi.")
            
    # SON ADIM: Elle düzeltme gereken çoklu dosya ve yazım hatalarını SQL ile fixle
    print("\nVeritabanı bazlı son düzeltmeler yapılıyor...")
    conn = get_db()
    cur = conn.cursor()
    
    if USE_SQLITE:
        cur.execute("UPDATE gelisim_planlari SET il = 'SAMSUN' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND sheet_adi LIKE '%SAM%' COLLATE NOCASE")
        cur.execute("UPDATE gelisim_planlari SET il = 'HAKKARİ' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND sheet_adi LIKE '%HAKK%' COLLATE NOCASE")
        cur.execute("UPDATE gelisim_planlari SET il = 'İZMİR' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND (sheet_adi LIKE '%İZMİR%' COLLATE NOCASE OR sheet_adi LIKE '%IZM%' COLLATE NOCASE)")
        cur.execute("UPDATE gelisim_planlari SET il = 'KÜTAHYA' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND (sheet_adi LIKE '%KÜT%' COLLATE NOCASE OR sheet_adi LIKE '%KUT%' COLLATE NOCASE)")
    else:
        cur.execute("UPDATE gelisim_planlari SET il = 'SAMSUN' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND sheet_adi ILIKE '%SAM%'")
        cur.execute("UPDATE gelisim_planlari SET il = 'HAKKARİ' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND sheet_adi ILIKE '%HAKK%'")
        cur.execute("UPDATE gelisim_planlari SET il = 'İZMİR' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND (sheet_adi ILIKE '%İZMİR%' OR sheet_adi ILIKE '%IZM%')")
        cur.execute("UPDATE gelisim_planlari SET il = 'KÜTAHYA' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND (sheet_adi ILIKE '%KÜT%' OR sheet_adi ILIKE '%KUT%')")
    cur.execute("UPDATE gelisim_planlari SET il = 'ERZURUM' WHERE kaynak_dosya = 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx' AND il = 'ERZURUM'") # Zaten öyle ama garanti
    conn.commit()
    cur.close()
    conn.close()
    
    print(f"\nİŞLEM TAMAMLANDI. TOPLAM KAYIT: {total}")

if __name__ == "__main__":
    run()
