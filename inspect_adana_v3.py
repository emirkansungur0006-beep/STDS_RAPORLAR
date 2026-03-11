import openpyxl
import os

path = r'C:\Users\EMİRKAN SUNGUR\Desktop\STDS_RAPORLAR\RAPORLAR\Gelişim_Planı_İller\ADANA.xlsx'

def inspect_adana():
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        print(f"\n===== SHEET: {sn} =====")
        ws = wb[sn]
        # İlk 30 satır, her hücreyi yazdır
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=30)):
            # None olmayan hücreleri filtreleyip göster
            cells = [str(c).replace('\n', ' ') for c in row if c is not None]
            if cells:
                print(f"L{i+1}: {' | '.join(cells)}")
    wb.close()

if __name__ == "__main__":
    inspect_adana()
