# -*- coding: utf-8 -*-
import os
import openpyxl
import re
from config import BASE_DIR
from reference_matcher import ReferenceMatcher, normalize_turkish

def clean_text(text):
    if text is None: return None
    t = str(text).strip()
    if t.lower() == 'none' or t == '': return None
    t = t.replace('_x000D_', ' ')
    t = re.sub(r'\s+', ' ', t)
    return t

class AdvancedGelisimParser:
    def __init__(self):
        self.matcher = ReferenceMatcher()
        self.keywords = {
            'kurum_hedefleri': ['hedef', 'kurum hedefleri'],
            'gerceklesme_suresi': ['süre', 'sure', 'zaman'],
            'mevcut_durum': ['mevcut durum', 'tespit'],
            'cozum_secenekleri': ['çözüm', 'cozum'],
            'etki_analizi': ['etki analizi'],
            'uygun_secenek': ['uygun', 'secenek'],
            'isbirligi_plani': ['işbirliği', 'isbirligi'],
            'uygulama_takvimi': ['takvim'],
            'hastane_col': ['hastane', 'tesis', 'kurum adı', 'kurum adi']
        }

    def detect_hospital(self, row, il=None):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not cells: return None
        
        row_str = " ".join(cells).upper()
        if 'ADI:' in row_str and ('SAĞLIK' in row_str or 'TESİS' in row_str or 'KURUM' in row_str):
            for ci, c_str in enumerate(cells):
                if 'ADI:' in c_str.upper():
                    potential = c_str.upper().split('ADI:')[-1].strip()
                    if len(potential) < 5 and ci + 1 < len(cells):
                        potential = cells[ci+1]
                    m = self.matcher.match_by_name(potential, il=il, threshold=0.7)
                    if not m: m = self.matcher.match_by_name(potential, threshold=0.7)
                    if m: return m

        if len(cells) > 1:
            potential_name = cells[0]
            potential_addr = cells[1]
            if len(potential_name) > 10 and (il and il.upper() in potential_addr.upper() or 'CAD.' in potential_addr.upper() or 'MAH.' in potential_addr.upper() or 'HASTANESİ' in potential_name.upper()):
                m = self.matcher.match_by_name(potential_name, il=il, threshold=0.75)
                if not m: m = self.matcher.match_by_name(potential_name, threshold=0.75)
                if m: return m
        
        if len(cells[0]) > 10:
             m = self.matcher.match_by_name(cells[0], il=il, threshold=0.85)
             if not m: m = self.matcher.match_by_name(cells[0], threshold=0.85)
             if m: return m
             
        return None

    def process_file(self, filepath):
        filename = os.path.basename(filepath)
        il_raw = os.path.splitext(filename)[0].split('-')[0].split('_')[0].strip().upper()
        if il_raw == 'SAMSUNKÜTAHYA': il_raw = 'SAMSUN'
        if 'ERZURUM' in il_raw and 'HAKKARİ' in il_raw: il_raw = 'ERZURUM'
        if il_raw == 'ISTANBUL': il_raw = 'İSTANBUL'
        
        il_match = self.matcher.match_by_name(il_raw, threshold=0.85)
        il_name = il_match['il'] if il_match else il_raw
        
        print(f"\n[{filename}] Default IL: {il_name}")
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        except Exception as e:
            print(f"    [!] Error: {e}")
            return []

        all_records = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            try:
                for r in ws.iter_rows(values_only=True, max_row=5000):
                    rows.append(r)
            except: continue
            
            if not rows: continue
            
            current_hosp = None
            header_map = None
            sheet_hosp = self.matcher.match_by_name(sheet_name, il=il_name, threshold=0.8)

            for i, row in enumerate(rows):
                if not any(row): continue
                filled_cells = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
                if not filled_cells: continue
                
                # Header detection
                if not header_map:
                    matches = {}
                    row_norm = [normalize_turkish(str(c)) if c else "" for c in row]
                    for col_db_name, kw_list in self.keywords.items():
                        for ci, val in enumerate(row_norm):
                            if any(kw == val or kw in val for kw in kw_list):
                                matches[col_db_name] = ci
                                break
                    if len(matches) >= 2 and ('kurum_hedefleri' in matches or 'mevcut_durum' in matches):
                        header_map = matches
                        continue

                # Hospital title detection (if few columns filled or no header)
                if not header_map or len(filled_cells) <= 3:
                    row_str = " ".join(filled_cells).upper()
                    if 'GELİŞİM PLAN' in row_str and 'HASTANE' not in row_str and 'TESİS' not in row_str:
                         pass
                    else:
                        new_hosp = self.detect_hospital(row, il=il_name)
                        if new_hosp:
                            current_hosp = new_hosp
                            header_map = None # Reset header
                            continue
                            
                # Data processing
                if header_map:
                    row_str = " ".join(filled_cells).upper()
                    if 'GELİŞİM PLAN' in row_str and len(filled_cells) < 3: continue
                    if 'HEDEF' in row_str and 'SÜRE' in row_str: continue

                    hastane_adi_row = None
                    if 'hastane_col' in header_map:
                        col_idx = header_map['hastane_col']
                        if col_idx < len(row):
                            h_val = row[col_idx]
                            if h_val:
                                h_match = self.matcher.match_by_name(str(h_val), il=il_name, threshold=0.7)
                                if not h_match: h_match = self.matcher.match_by_name(str(h_val), threshold=0.7)
                                if h_match:
                                    hastane_adi_row = h_match
                                else:
                                    hastane_adi_row = {'il': il_name, 'kurum_adi': str(h_val).strip(), 'kurum_kodu': None}

                    hastane_match = hastane_adi_row or current_hosp or sheet_hosp
                    
                    if hastane_match:
                        hastane_adi = hastane_match.get('kurum_adi', f"{il_name} - {sheet_name.upper()}")
                        kurum_kodu = hastane_match.get('kurum_kodu')
                        gercek_il = hastane_match.get('il', il_name)
                    else:
                        hastane_adi = f"{il_name} - {sheet_name.upper()}"
                        kurum_kodu = None
                        gercek_il = il_name
                        
                    rec = {
                        'il': gercek_il,
                        'hastane_adi': hastane_adi,
                        'kurum_kodu': kurum_kodu,
                        'kaynak_dosya': filename,
                        'sheet_adi': sheet_name
                    }
                    
                    for col_db_name, ci in header_map.items():
                        if col_db_name == 'hastane_col': continue
                        val = row[ci] if ci < len(row) else None
                        rec[col_db_name] = clean_text(val)
                    
                    if not rec.get('kurum_hedefleri') or len(str(rec.get('kurum_hedefleri'))) < 5:
                        continue

                    all_records.append(rec)

        wb.close()
        return all_records

if __name__ == '__main__':
    parser = AdvancedGelisimParser()
    test_files = [
        os.path.join(BASE_DIR, 'RAPORLAR', 'Gelişim_Planı_İller', 'GAZİANTEP.xlsx'),
        os.path.join(BASE_DIR, 'RAPORLAR', 'Gelişim_Planı_İller', 'SAMSUNKÜTAHYA.xlsx'),
        os.path.join(BASE_DIR, 'RAPORLAR', 'Gelişim_Planı_İller', 'İSTANBUL KHHB-2 2026 YILI HASTANE GELİŞİM PLANLARI (YAN YANA ŞEKLİNDE).xlsx'),
        os.path.join(BASE_DIR, 'RAPORLAR', 'Gelişim_Planı_İller', 'ERZURUM HAKKARİ İZMİR KÜYAHYA SAMSUN.xlsx'),
    ]
    
    for f in test_files:
        if os.path.exists(f):
            print(f"Testing {os.path.basename(f)}...")
            recs = parser.process_file(f)
            print(f"Found {len(recs)} records.")
            if recs:
                print("First record:", {k: v for k, v in list(recs[0].items())[:6]})
                print("Last record:", {k: v for k, v in list(recs[-1].items())[:6]})
                ils = set(r['il'] for r in recs)
                print(f"Found Provinces: {ils}\n")

