import openpyxl
path = r'C:\Users\EMİRKAN SUNGUR\Desktop\STDS_RAPORLAR\RAPORLAR\Gelişim_Planı_İller\ADANA.xlsx'
wb = openpyxl.load_workbook(path, read_only=True)
print(f"Sheet names ({len(wb.sheetnames)}):")
print(wb.sheetnames)
wb.close()
