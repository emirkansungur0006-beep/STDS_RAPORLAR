import openpyxl
import os

path = r'C:\Users\EMİRKAN SUNGUR\Desktop\STDS_RAPORLAR\RAPORLAR\Gelişim_Planı_İller\ADANA.xlsx'

def search_adana():
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    print(f"Total rows in {ws.title}: {ws.max_row}")
    # İlk 500 satırda gezelim ve tesis adı içerenleri yazdıralım
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=500)):
        row_str = " | ".join([str(c) for c in row if c is not None])
        if "SAĞLIK TESİSİ" in row_str.upper() or "KURUM ADI" in row_str.upper() or "KURUM KODU" in row_str.upper():
            print(f"L{i+1}: {row_str}")
        # Hangi satırlarda veri yoğunluğu var?
        if "HEDEF" in row_str.upper() and "MEVCUT DURUM" in row_str.upper():
             print(f"--- DATA HEADER AT L{i+1} ---")
    wb.close()

if __name__ == "__main__":
    search_adana()
