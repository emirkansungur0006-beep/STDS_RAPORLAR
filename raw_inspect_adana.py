import openpyxl
path = r'C:\Users\EMİRKAN SUNGUR\Desktop\STDS_RAPORLAR\RAPORLAR\Gelişim_Planı_İller\ADANA.xlsx'
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb.active
print(f"--- ROW 7 ---")
r7 = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
print(f"R7 TYPE: {type(r7)}")
print(f"R7 CONTENT: {r7}")
print(f"--- ROW 261 ---")
r261 = list(ws.iter_rows(min_row=261, max_row=261, values_only=True))[0]
print(f"R261 CONTENT: {r261}")
wb.close()
