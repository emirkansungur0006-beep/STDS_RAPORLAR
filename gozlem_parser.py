# -*- coding: utf-8 -*-
"""
STDS Raporlar - Gözlem Formları Parser
XLSX gözlem formlarını okur ve veritabanına yazar.
"""
import os
import re
import sqlite3
import openpyxl
from config import GOZLEM_DIR
from app import get_db
from reference_matcher import ReferenceMatcher, normalize_turkish


def extract_hospital_name_from_filename(filename):
    """Dosya adından hastane adını çıkar"""
    # "T.C. SAĞLIK BAKANLIĞI XXX HASTANESİ.xlsx" formatı
    name = os.path.splitext(filename)[0]
    name = re.sub(r'^T\.?C\.?\s*SAĞLIK\s*BAKANLIĞI\s*', '', name, flags=re.IGNORECASE)
    return name.strip()


def parse_gozlem_xlsx(filepath, il, matcher):
    """Bir gözlem XLSX dosyasını parse et"""
    filename = os.path.basename(filepath)
    hastane_adi_raw = extract_hospital_name_from_filename(filename)

    # Referans eşleştirme
    hospital = matcher.match_by_name(hastane_adi_raw, il=il, threshold=0.88)
    
    hastane_adi_exact = filename.replace('.xlsx', '').replace('.xls', '').strip()
    
    # Görsellerle birebir eşleşmesi için dosya adını kullanıyoruz:
    hastane_adi = hastane_adi_exact
    
    if hospital:
        kurum_kodu = hospital['kurum_kodu']
        ilce = hospital['ilce']
    else:
        kurum_kodu = None
        ilce = None

    records = []

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [HATA] Dosya açılamadı {filename}: {e}")
        return records

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Header satırını bul
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows):
            row_str = ' '.join([str(c).lower() if c else '' for c in row])
            if 'soru' in row_str and ('derece' in row_str or 'not' in row_str):
                header_row = row
                header_idx = i
                break

        if header_row is None:
            # İlk satırı header olarak varsay
            header_row = rows[0]
            header_idx = 0

        # Sütun mapping
        col_map = {}
        for ci, cell in enumerate(header_row):
            if cell:
                cell_str = normalize_turkish(str(cell))
                if 'bölüm' in cell_str or 'bolum' in cell_str:
                    col_map['bolum'] = ci
                elif 'soru no' in cell_str or 'soru_no' in cell_str or cell_str == 'no':
                    col_map['soru_no'] = ci
                elif 'soru' in cell_str and 'no' not in cell_str:
                    col_map['soru'] = ci
                elif 'derece' in cell_str or 'puan' in cell_str:
                    col_map['derece'] = ci
                elif 'not' in cell_str or 'açıklama' in cell_str:
                    col_map['notlar'] = ci

        # Veri satırlarını oku
        current_bolum = sheet_name
        for row in rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            bolum = str(row[col_map['bolum']]).strip() if 'bolum' in col_map and row[col_map['bolum']] else current_bolum
            if bolum and bolum != 'None':
                current_bolum = bolum

            soru_no = None
            if 'soru_no' in col_map and row[col_map['soru_no']]:
                try:
                    soru_no = int(row[col_map['soru_no']])
                except (ValueError, TypeError):
                    pass

            soru = str(row[col_map['soru']]).strip() if 'soru' in col_map and row[col_map['soru']] else None
            derece = str(row[col_map['derece']]).strip() if 'derece' in col_map and row[col_map['derece']] else None
            notlar = str(row[col_map['notlar']]).strip() if 'notlar' in col_map and row[col_map['notlar']] else None

            if not soru and not derece:
                continue

            if derece == 'None':
                derece = None
            if notlar == 'None':
                notlar = None

            records.append({
                'kurum_kodu': kurum_kodu,
                'il': il,
                'ilce': ilce,
                'hastane_adi': hastane_adi,
                'bolum': current_bolum,
                'soru_no': soru_no,
                'soru': soru,
                'verilen_derece': derece,
                'notlar': notlar,
                'kaynak_dosya': filename,
                'sheet_adi': sheet_name
            })

    wb.close()
    return records


def save_gozlem_records(records):
    """Gözlem kayıtlarını veritabanına yaz"""
    if not records:
        return 0

    conn = get_db()
    cur = conn.cursor()
    
    from config import USE_SQLITE
    placeholders = '?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?' if USE_SQLITE else '%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s'
    
    count = 0

    for rec in records:
        try:
            cur.execute(f"""
                INSERT INTO gozlem_formlari
                (kurum_kodu, il, ilce, hastane_adi, bolum, soru_no, soru,
                 verilen_derece, notlar, kaynak_dosya, sheet_adi)
                VALUES ({placeholders})
            """, (
                rec['kurum_kodu'], rec['il'], rec['ilce'], rec['hastane_adi'],
                rec['bolum'], rec['soru_no'], rec['soru'],
                rec['verilen_derece'], rec['notlar'], rec['kaynak_dosya'],
                rec['sheet_adi']
            ))
            count += 1
        except Exception as e:
            print(f"  [HATA] Kayıt yazılamadı: {e}")
            conn.rollback()
            # conn ve cur nesnelerini get_db üzerinden yenilemeye gerek yok, 
            # sadece hatayı loglayıp devam edelim.

    conn.commit()
    cur.close()
    conn.close()
    return count


def process_all_gozlem():
    """Tüm gözlem formlarını işle"""
    print("=" * 60)
    print("GÖZLEM FORMLARI İŞLENİYOR")
    print("=" * 60)

    matcher = ReferenceMatcher()
    total_records = 0
    total_files = 0

    # Mevcut gözlem verilerini temizle
    conn = get_db()
    cur = conn.cursor()
    from config import USE_SQLITE
    if USE_SQLITE:
        cur.execute("DELETE FROM gozlem_formlari")
    else:
        cur.execute("TRUNCATE TABLE gozlem_formlari")
    conn.commit()
    cur.close()
    conn.close()

    for il_folder in sorted(os.listdir(GOZLEM_DIR)):
        il_path = os.path.join(GOZLEM_DIR, il_folder)
        if not os.path.isdir(il_path):
            continue

        il_name = il_folder.strip()
        print(f"\n[İl: {il_name}]")

        for filename in sorted(os.listdir(il_path)):
            if not filename.lower().endswith(('.xlsx', '.xls')):
                continue

            filepath = os.path.join(il_path, filename)
            print(f"  İşleniyor: {filename}")

            records = parse_gozlem_xlsx(filepath, il_name, matcher)
            if records:
                saved = save_gozlem_records(records)
                total_records += saved
                total_files += 1
                print(f"    -> {saved} kayıt eklendi")
            else:
                print(f"    -> Kayıt bulunamadı")

    print(f"\n{'=' * 60}")
    print(f"TOPLAM: {total_files} dosya, {total_records} kayıt işlendi.")
    print(f"{'=' * 60}")
    return total_records


if __name__ == '__main__':
    process_all_gozlem()
