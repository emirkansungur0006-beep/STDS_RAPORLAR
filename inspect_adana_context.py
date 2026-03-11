import openpyxl
import os

path = r'C:\Users\EMİRKAN SUNGUR\Desktop\STDS_RAPORLAR\RAPORLAR\Gelişim_Planı_İller\ADANA.xlsx'

def inspect_adana_context():
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    
    # L1-L15 arasını görelim
    print("--- L1 to L15 ---")
    for i, row in enumerate(ws.iter_rows(values_only=True, min_row=1, max_row=15)):
        print(f"L{i+1}: {row}")
        
    # L260-L275 arasını görelim
    print("\n--- L260 to L275 ---")
    for i, row in enumerate(ws.iter_rows(values_only=True, min_row=260, max_row=275)):
        print(f"L{i+260}: {row}")
    
    wb.close()

if __name__ == "__main__":
    inspect_adana_context()
