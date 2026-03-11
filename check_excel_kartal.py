import openpyxl
filepath = r"C:\Users\EMİRKAN SUNGUR\Desktop\STDS_RAPORLAR\RAPORLAR\GÖZLEM FORMLARI\İSTANBUL\T.C. SAĞLIK BAKANLIĞI KARTAL AĞIZ VE DİŞ SAĞLIĞI MERKEZİ.xlsx"
wb = openpyxl.load_workbook(filepath, read_only=True)
sheet = wb[wb.sheetnames[0]]
print("--- ROW 2 CELLS ---")
for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=5, values_only=True)):
    print(f"Row {i+2}: Col B: {row[1]}, Col C: {row[2]}")
